from io import BytesIO, StringIO
from typing import Tuple
from PyPDF2 import PdfReader
from docx import Document

from . import ocr as _ocr_module

# Accepted import extensions, surfaced to callers/UI so the list stays in one place.
SUPPORTED_EXTENSIONS = ('.txt', '.csv', '.pdf', '.docx', '.xlsx')

# Tags wrapping a table extracted from a document. Downstream chunking keeps the
# block between these markers intact so a Schedule-of-Events matrix is never
# split across chunks. Kept here so ingestion and docstore agree on one spelling.
TABLE_OPEN = '[[TABLE]]'
TABLE_CLOSE = '[[/TABLE]]'


def _extract_pdf_with_pdfplumber(file_bytes: bytes) -> str:
    """Structure-aware PDF extraction: page text plus tables as tagged TSV blocks.

    pdfplumber preserves table grid structure that PyPDF2's flat ``extract_text``
    destroys — essential for nested tables and Schedule-of-Events matrices. Each
    detected table is emitted as a tab-delimited block wrapped in
    ``TABLE_OPEN``/``TABLE_CLOSE`` so it survives chunking as one unit.

    Imported lazily so pdfplumber stays an optional dependency: callers fall back
    to PyPDF2 when it (or this extraction) is unavailable.
    """
    import pdfplumber  # lazy: optional dependency

    parts: list[str] = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ''
            if page_text.strip():
                parts.append(page_text)
            for table in page.find_tables():
                rows = table.extract() or []
                lines = []
                for row in rows:
                    cells = ['' if c is None else str(c).replace('\n', ' ').strip()
                             for c in row]
                    if any(cells):
                        lines.append('\t'.join(cells))
                if lines:
                    parts.append(f'{TABLE_OPEN}\n' + '\n'.join(lines)
                                 + f'\n{TABLE_CLOSE}')
    return '\n'.join(parts)


def extract_text_from_pdf_bytes(file_bytes: bytes) -> str:
    """Extract text from a PDF, preferring structure-aware extraction.

    Fallback chain: pdfplumber (tables preserved) -> PyPDF2 (flat text) -> OCR
    (scanned/image PDFs). Each stage degrades gracefully when its dependency is
    absent or yields nothing.
    """
    text = ''
    try:
        text = _extract_pdf_with_pdfplumber(file_bytes)
    except Exception:
        text = ''

    if not text.strip():
        # pdfplumber unavailable or produced nothing — flat PyPDF2 extraction.
        try:
            reader = PdfReader(BytesIO(file_bytes))
            text = '\n'.join(page.extract_text() or '' for page in reader.pages)
        except Exception:
            text = ''

    if not text.strip():
        # PDF had no selectable text (e.g. scanned image) — try OCR fallback.
        # ocr.ocr_pdf_bytes degrades gracefully when dependencies are absent.
        text = _ocr_module.ocr_pdf_bytes(file_bytes)
    return text


def extract_text_from_docx_bytes(file_bytes: bytes) -> str:
    doc = Document(BytesIO(file_bytes))
    return '\n'.join(p.text for p in doc.paragraphs)


def extract_text_from_xlsx_bytes(file_bytes: bytes) -> str:
    """Flatten every sheet into tab-separated rows so the parser sees the cells."""
    from openpyxl import load_workbook  # lazy: keep import cost off the hot path
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        if len(wb.worksheets) > 1:
            lines.append(f'# Sheet: {ws.title}')
        for row in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c) for c in row]
            if any(cell.strip() for cell in cells):
                lines.append('\t'.join(cells))
    wb.close()
    return '\n'.join(lines)


def ingest_uploaded_file(filename: str, file_bytes: bytes) -> Tuple[str, str]:
    lower = filename.lower()
    if lower.endswith('.txt'):
        return file_bytes.decode('utf-8', errors='ignore'), 'txt'
    if lower.endswith('.csv'):
        return file_bytes.decode('utf-8-sig', errors='ignore'), 'csv'
    if lower.endswith('.pdf'):
        return extract_text_from_pdf_bytes(file_bytes), 'pdf'
    if lower.endswith('.docx'):
        return extract_text_from_docx_bytes(file_bytes), 'docx'
    if lower.endswith('.xlsx'):
        return extract_text_from_xlsx_bytes(file_bytes), 'xlsx'
    if lower.endswith('.doc') or lower.endswith('.xls'):
        raise ValueError(
            'Legacy .doc/.xls is not supported. Please re-save as .docx or .xlsx.'
        )
    raise ValueError(
        'Unsupported file type. Use ' + ', '.join(SUPPORTED_EXTENSIONS)
    )
